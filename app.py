import asyncio
import json
import os
import re
import math
import html
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

import uvicorn
from docx import Document
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.middleware import SlowAPIMiddleware
from google import genai
from google.genai import types
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator
from pypdf import PdfReader

# Carga estricta de variables de entorno
load_dotenv()

app = FastAPI(title="Central de IA Premium - Backend Pro")
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_middleware(SlowAPIMiddleware)

# --- CONFIGURACIÓN DE CORS ---
CORS_ALLOW_ORIGINS = [
    origin.strip()
    for origin in os.getenv("CORS_ALLOW_ORIGINS", "*").split(",")
    if origin.strip()
]
CORS_ALLOW_CREDENTIALS = "*" not in CORS_ALLOW_ORIGINS
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOW_ORIGINS,
    allow_credentials=CORS_ALLOW_CREDENTIALS,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- VERIFICACIÓN Y ASIGNACIÓN DE API KEY ---
API_KEY = os.getenv("GEMINI_API_KEY")
client: Optional[genai.Client] = None
if API_KEY and API_KEY != "AIzaSyTuClaveRealDeGoogleAqui":
    client = genai.Client(api_key=API_KEY)

# Executor para correr el streaming síncrono de Gemini sin bloquear el event loop
_executor = ThreadPoolExecutor(max_workers=4)

# --- CONFIGURACIÓN DE PERSONALIDADES (SYSTEM PROMPTS) ---
personalidades = {
    "ml_engineer": (
        "Eres un Ingeniero experto en Machine Learning y arquitecturas de Deep Learning. "
        "Tus respuestas deben ser técnicas, precisas y enfocadas en optimización de código, "
        "entrenamiento de modelos y despliegue eficiente de infraestructura."
    ),
    "profesional": (
        "Eres un Ingeniero de Audio de elite y Diseñador de Circuitos Electrónicos. "
        "Tienes un conocimiento profundo de parámetros Thiele-Small, alineamientos de cajas acústicas, "
        "procesadores digitales (DSP) y diseño de filtros analógicos. Cuando te soliciten cálculos de filtros "
        "o crossovers, menciona siempre de forma clara la frecuencia de corte elegida con el formato 'XXXX Hz' "
        "para que el sistema visual pueda renderizar el simulador gráfico."
    ),
    "creativo": "Eres un asistente general de IA, útil, directo, sumamente inteligente y de mente abierta.",
    "educador": "Eres un tutor académico riguroso pero didáctico. Explicas conceptos complejos desglosándolos paso a paso.",
    "copywriter": "Eres un redactor creativo experto en SEO, copywriting y estrategias de contenido técnico de alto impacto.",
}

# --- MODELOS DE DATOS DE ENTRADA (PYDANTIC) ---
class LoginRequest(BaseModel):
    usuario: str = Field(..., min_length=1, max_length=100)
    password: str = Field(..., min_length=1, max_length=200)


class ChatRequest(BaseModel):
    usuario: str = Field(..., min_length=1, max_length=100)
    id_chat: str = Field(..., min_length=1, max_length=100)
    personalidad: str = Field(..., min_length=1, max_length=50)
    mensaje: str = Field(..., max_length=10000)  # límite de seguridad
    modelo_elegido: str = "gemini-2.5-flash-lite"


class MemoriaRequest(BaseModel):
    memoria: str = Field(default="", max_length=20000)


class RenombrarChatRequest(BaseModel):
    nuevo_nombre: str = Field(..., min_length=1, max_length=100)


class AdminUsuarioRequest(BaseModel):
    password: str = Field(..., min_length=4, max_length=200)
    rol: str = Field(default="usuario", pattern="^(admin|usuario)$")
    activo: bool = True


class AudioToolRequest(BaseModel):
    herramienta: str = Field(..., pattern="^(crossover|caja_sellada|impedancia_paralelo)$")
    frecuencia_hz: Optional[float] = Field(default=None, gt=0, le=100000)
    impedancia_ohm: Optional[float] = Field(default=None, gt=0, le=100000)
    orden: int = Field(default=1, ge=1, le=2)
    volumen_litros: Optional[float] = Field(default=None, gt=0, le=10000)
    vas_litros: Optional[float] = Field(default=None, gt=0, le=10000)
    qts: Optional[float] = Field(default=None, gt=0, le=10)
    impedancias_ohm: Optional[List[float]] = None


class AmpRequest(BaseModel):
    topologia: str = Field(..., min_length=1, max_length=50)  # "clase_a", "clase_ab", "common_emitter"
    vcc: float = Field(..., gt=0, le=1000)              # Tensión de alimentación (V)
    potencia_w: float = Field(..., ge=0, le=100000)       # Potencia de salida deseada (W)
    impedancia_carga: float = Field(..., gt=0, le=100000) # Impedancia del parlante (Ω)
    transistor: str = Field(default="2N3055", min_length=1, max_length=80)  # Transistor sugerido

    @field_validator("topologia")
    @classmethod
    def normalizar_topologia(cls, valor: str) -> str:
        return valor.strip().lower()


# --- CONFIGURACIÓN DE MODELOS ---
MODELOS_DISPONIBLES = ["gemini-2.5-flash-lite", "gemini-2.5-flash"]

# Máximo de mensajes de contexto enviados al modelo
LIMITE_CONTEXTO = 30
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# --- PERSISTENCIA LOCAL (CARPETA DE HISTORIALES) ---
RUTA_BASE = Path(__file__).resolve().parent
RUTA_HISTORIALES = RUTA_BASE / "historiales"
RUTA_HISTORIALES.mkdir(exist_ok=True)
IDENTIFICADOR_SEGURO = re.compile(r"[^a-zA-Z0-9_-]")
_historial_lock = Lock()


def normalizar_identificador(valor: str, campo: str) -> str:
    limpio = IDENTIFICADOR_SEGURO.sub("_", valor.strip())
    if not limpio:
        raise HTTPException(status_code=400, detail=f"El campo {campo} no puede estar vacío")
    return limpio


def obtener_ruta_json(usuario: str, id_chat: str) -> Path:
    usuario_seguro = normalizar_identificador(usuario, "usuario")
    chat_seguro = normalizar_identificador(id_chat, "id_chat")
    ruta = (RUTA_HISTORIALES / f"{usuario_seguro}_{chat_seguro}.json").resolve()
    if RUTA_HISTORIALES.resolve() not in ruta.parents:
        raise HTTPException(status_code=400, detail="Ruta de historial inválida")
    return ruta


def cargar_historial_local(usuario: str, id_chat: str) -> List[Dict[str, str]]:
    ruta = obtener_ruta_json(usuario, id_chat)
    if ruta.exists():
        try:
            with ruta.open("r", encoding="utf-8") as f:
                historial = json.load(f)
            if isinstance(historial, list):
                return historial
        except (json.JSONDecodeError, OSError):
            return []
    return []


def guardar_en_json_local(usuario: str, id_chat: str, mensaje_user: str, respuesta_bot: str) -> None:
    ruta = obtener_ruta_json(usuario, id_chat)
    with _historial_lock:
        historial = cargar_historial_local(usuario, id_chat)
        historial.append({"role": "user", "content": mensaje_user})
        historial.append({"role": "assistant", "content": respuesta_bot})
        ruta_temporal = ruta.with_suffix(".json.tmp")
        with ruta_temporal.open("w", encoding="utf-8") as f:
            json.dump(historial, f, ensure_ascii=False, indent=4)
        ruta_temporal.replace(ruta)


def formatear_sse_data(texto: str) -> str:
    """Formatea texto arbitrario como evento SSE válido, terminando con doble salto de línea."""
    texto = texto.replace("\r", "")
    lineas = texto.split("\n")
    return "".join(f"data: {linea}\n" for linea in lineas) + "\n\n"


def obtener_ruta_usuarios() -> Path:
    return RUTA_BASE / "usuarios.json"


def cargar_config_usuarios() -> Dict[str, Dict[str, Any]]:
    ruta_usuarios = obtener_ruta_usuarios()
    if not ruta_usuarios.exists():
        return {}
    try:
        data = json.loads(ruta_usuarios.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

    usuarios: Dict[str, Dict[str, Any]] = {}
    for usuario, config in data.items():
        if isinstance(config, dict) and isinstance(config.get("password"), str):
            usuarios[usuario] = {
                "password": config["password"],
                "rol": config.get("rol", "usuario"),
                "activo": bool(config.get("activo", True)),
            }
    return usuarios


def guardar_config_usuarios(usuarios: Dict[str, Dict[str, Any]]) -> None:
    ruta_usuarios = obtener_ruta_usuarios()
    ruta_temporal = ruta_usuarios.with_suffix(".json.tmp")
    with ruta_temporal.open("w", encoding="utf-8") as f:
        json.dump(usuarios, f, ensure_ascii=False, indent=2)
    ruta_temporal.replace(ruta_usuarios)


def cargar_usuarios_validos() -> Dict[str, str]:
    return {
        usuario: config["password"]
        for usuario, config in cargar_config_usuarios().items()
        if config.get("activo", True)
    }


def validar_admin(usuario: str) -> None:
    config = cargar_config_usuarios().get(usuario)
    if not config or config.get("rol") != "admin" or not config.get("activo", True):
        raise HTTPException(status_code=403, detail="Se requiere un usuario administrador activo")


# --- FUNCIONES DE MEMORIA ---
def obtener_ruta_memoria(usuario: str) -> Path:
    usuario_seguro = normalizar_identificador(usuario, "usuario")
    return RUTA_HISTORIALES / f"{usuario_seguro}_memoria.txt"


def cargar_memoria_local(usuario: str) -> str:
    ruta = obtener_ruta_memoria(usuario)
    if ruta.exists():
        return ruta.read_text(encoding="utf-8")
    return ""


def guardar_memoria_local(usuario: str, memoria: str) -> None:
    ruta = obtener_ruta_memoria(usuario)
    ruta.write_text(memoria, encoding="utf-8")


# --- ENDPOINTS ---

# 1. RUTA RAÍZ: Sirve el frontend
@app.get("/", response_class=HTMLResponse)
async def serve_index():
    ruta_index = RUTA_BASE / "templates" / "index.html"
    if ruta_index.exists():
        return ruta_index.read_text(encoding="utf-8")
    raise HTTPException(status_code=404, detail="No se encontró el index.html dentro de la carpeta templates")


# 2. LOGIN
@app.post("/login")
async def login(req: LoginRequest):
    usuarios_config = cargar_config_usuarios()
    config = usuarios_config.get(req.usuario)
    if config and config.get("activo", True) and config.get("password") == req.password:
        return {"status": "success", "usuario": req.usuario, "rol": config.get("rol", "usuario")}
    raise HTTPException(status_code=401, detail="Credenciales incorrectas")


# 3. ADMINISTRACIÓN DE USUARIOS
@app.get("/admin/usuarios")
async def listar_usuarios_admin(admin: str = Query(...)):
    validar_admin(admin)
    usuarios = cargar_config_usuarios()
    return {
        "usuarios": [
            {"usuario": usuario, "rol": config.get("rol", "usuario"), "activo": config.get("activo", True)}
            for usuario, config in sorted(usuarios.items())
        ]
    }


@app.put("/admin/usuarios/{usuario}")
async def guardar_usuario_admin(usuario: str, req: AdminUsuarioRequest, admin: str = Query(...)):
    validar_admin(admin)
    usuario_seguro = normalizar_identificador(usuario, "usuario")
    usuarios = cargar_config_usuarios()
    usuarios[usuario_seguro] = {"password": req.password, "rol": req.rol, "activo": req.activo}
    guardar_config_usuarios(usuarios)
    return {"status": "success", "usuario": usuario_seguro, "rol": req.rol, "activo": req.activo}


@app.delete("/admin/usuarios/{usuario}")
async def borrar_usuario_admin(usuario: str, admin: str = Query(...)):
    validar_admin(admin)
    usuario_seguro = normalizar_identificador(usuario, "usuario")
    if usuario_seguro == admin:
        raise HTTPException(status_code=400, detail="No podés borrar tu propio usuario administrador")
    usuarios = cargar_config_usuarios()
    if usuario_seguro not in usuarios:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    del usuarios[usuario_seguro]
    guardar_config_usuarios(usuarios)
    return {"status": "success"}


# 4. MODELOS
@app.get("/modelos")
async def listar_modelos():
    return {"modelos": MODELOS_DISPONIBLES}


# 4. MEMORIA
@app.get("/memoria/{usuario}")
async def obtener_memoria(usuario: str):
    return {"memoria": cargar_memoria_local(usuario)}


@app.put("/memoria/{usuario}")
async def guardar_memoria(usuario: str, req: MemoriaRequest):
    guardar_memoria_local(usuario, req.memoria)
    return {"status": "success", "memoria": req.memoria.strip()}


@app.delete("/memoria/{usuario}")
async def borrar_memoria(usuario: str):
    ruta = obtener_ruta_memoria(usuario)
    if ruta.exists():
        ruta.unlink()
    return {"status": "success"}


# 5. CHATS — listar con búsqueda opcional
@app.get("/chats/{usuario}")
async def listar_chats(usuario: str, q: Optional[str] = Query(default=None)):
    usuario_seguro = normalizar_identificador(usuario, "usuario")
    prefijo = f"{usuario_seguro}_"
    chats_usuario = []

    for ruta in RUTA_HISTORIALES.glob(f"{prefijo}*.json"):
        nombre_chat = ruta.stem.removeprefix(prefijo)
        if nombre_chat == "memoria":
            continue
        if q:
            termino = q.lower()
            if termino in nombre_chat.lower():
                chats_usuario.append(nombre_chat)
                continue
            try:
                historial = cargar_historial_local(usuario, nombre_chat)
                contenido_completo = " ".join(m.get("content", "") for m in historial).lower()
                if termino in contenido_completo:
                    chats_usuario.append(nombre_chat)
            except Exception:
                pass
        else:
            chats_usuario.append(nombre_chat)

    return {"chats": sorted(chats_usuario, reverse=True)}


# 6. CHATS — cargar historial de un chat específico
@app.get("/chats/{usuario}/{id_chat}")
async def obtener_chat(usuario: str, id_chat: str):
    historial = cargar_historial_local(usuario, id_chat)
    return {"mensajes": historial}


# 7. CHATS — borrar un chat
@app.delete("/chats/{usuario}/{id_chat}")
async def borrar_chat(usuario: str, id_chat: str):
    ruta = obtener_ruta_json(usuario, id_chat)
    if not ruta.exists():
        raise HTTPException(status_code=404, detail="Chat no encontrado")
    ruta.unlink()
    return {"status": "success"}


# 8. CHATS — renombrar un chat
@app.put("/chats/{usuario}/{id_chat}/renombrar")
async def renombrar_chat(usuario: str, id_chat: str, req: RenombrarChatRequest):
    ruta_original = obtener_ruta_json(usuario, id_chat)
    if not ruta_original.exists():
        raise HTTPException(status_code=404, detail="Chat no encontrado")

    nuevo_nombre_limpio = normalizar_identificador(req.nuevo_nombre, "nuevo_nombre")
    ruta_nueva = obtener_ruta_json(usuario, nuevo_nombre_limpio)

    if ruta_nueva.exists():
        raise HTTPException(status_code=409, detail="Ya existe un chat con ese nombre")

    ruta_original.rename(ruta_nueva)
    return {"status": "success", "id_chat": nuevo_nombre_limpio}


# 9. RUTA PÚBLICA COMPARTIDA
@app.get("/share/{usuario}/{id_chat}", response_class=HTMLResponse)
async def share_chat(usuario: str, id_chat: str):
    historial = cargar_historial_local(usuario, id_chat)
    if not historial:
        raise HTTPException(status_code=404, detail="Conversación no encontrada o vacía")

    filas_html = ""
    for m in historial:
        rol = m.get("role", "")
        contenido = html.escape(str(m.get("content", "")))
        label = "Tú" if rol == "user" else "IA"
        color = "#3b82f6" if rol == "user" else "#10b981"
        filas_html += f'<p><strong style="color:{color}">{label}:</strong> {contenido}</p><hr>'

    return f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Chat compartido</title>
<style>body{{font-family:sans-serif;max-width:800px;margin:40px auto;padding:20px;background:#111;color:#eee}}hr{{border-color:#333}}</style>
</head><body><h2>Conversación compartida</h2>{filas_html}</body></html>"""


# 10. MOTOR DE CHAT CON STREAMING REAL
@app.post("/chat")
async def chat(req: ChatRequest):
    if client is None:
        raise HTTPException(
            status_code=503,
            detail="No se encontró la variable GEMINI_API_KEY configurada en el entorno",
        )

    if req.modelo_elegido not in MODELOS_DISPONIBLES:
        raise HTTPException(status_code=400, detail="Modelo no soportado")

    historial_completo = cargar_historial_local(req.usuario, req.id_chat)
    historial_recortado = historial_completo[-LIMITE_CONTEXTO:]

    contents = []
    for m in historial_recortado:
        if not isinstance(m, dict) or "role" not in m or "content" not in m:
            continue
        role_google = "model" if m["role"] == "assistant" else "user"
        contents.append(
            types.Content(
                role=role_google,
                parts=[types.Part.from_text(text=str(m["content"]))],
            )
        )
    contents.append(types.Content(role="user", parts=[types.Part.from_text(text=req.mensaje)]))

    instruccion_sistema = personalidades.get(req.personalidad, "Eres un asistente útil.")
    memoria_usuario = cargar_memoria_local(req.usuario)
    if memoria_usuario:
        instruccion_sistema += (
            "\n\nMemoria permanente del usuario. Úsala solo cuando sea relevante y no la reveles "
            f"literalmente salvo que el usuario lo pida:\n{memoria_usuario}"
        )

    queue: asyncio.Queue = asyncio.Queue()
    loop = asyncio.get_running_loop()

    def _llamar_gemini_en_hilo():
        try:
            response_stream = client.models.generate_content_stream(
                model=req.modelo_elegido,
                contents=contents,
                config=types.GenerateContentConfig(
                    system_instruction=instruccion_sistema,
                    temperature=0.7,
                ),
            )
            for chunk in response_stream:
                if chunk.text:
                    loop.call_soon_threadsafe(queue.put_nowait, chunk.text)
        except Exception as e:
            loop.call_soon_threadsafe(queue.put_nowait, f"\n[Error en streaming: {str(e)}]")
        finally:
            loop.call_soon_threadsafe(queue.put_nowait, None)

    async def generador_streaming():
        texto_completo_bot = ""
        loop.run_in_executor(_executor, _llamar_gemini_en_hilo)

        while True:
            chunk_text = await queue.get()
            if chunk_text is None:
                break

            texto_completo_bot += chunk_text
            yield formatear_sse_data(chunk_text)
            await asyncio.sleep(0)

        if texto_completo_bot:
            guardar_en_json_local(req.usuario, req.id_chat, req.mensaje, texto_completo_bot)

    return StreamingResponse(
        generador_streaming(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# 11. SUBIDA Y EXTRACCIÓN DE ARCHIVOS
@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="El archivo debe tener nombre")

    nombre_archivo = file.filename.lower()

    try:
        if nombre_archivo.endswith((".txt", ".csv", ".json", ".py", ".html", ".css", ".md")):
            bytes_archivo = await file.read(MAX_UPLOAD_BYTES + 1)
            if len(bytes_archivo) > MAX_UPLOAD_BYTES:
                raise HTTPException(status_code=413, detail="El archivo supera el límite de 10 MB")
            contenido_extraido = bytes_archivo.decode("utf-8", errors="ignore")
        elif nombre_archivo.endswith(".pdf"):
            data = await file.read(MAX_UPLOAD_BYTES + 1)
            if len(data) > MAX_UPLOAD_BYTES:
                raise HTTPException(status_code=413, detail="El archivo supera el límite de 10 MB")
            pdf = PdfReader(BytesIO(data))
            contenido_extraido = "\n".join([(p.extract_text() or "") for p in pdf.pages])
        elif nombre_archivo.endswith(".docx"):
            data = await file.read(MAX_UPLOAD_BYTES + 1)
            if len(data) > MAX_UPLOAD_BYTES:
                raise HTTPException(status_code=413, detail="El archivo supera el límite de 10 MB")
            doc = Document(BytesIO(data))
            contenido_extraido = "\n".join(p.text for p in doc.paragraphs)
        elif nombre_archivo.endswith(".xlsx"):
            data = await file.read(MAX_UPLOAD_BYTES + 1)
            if len(data) > MAX_UPLOAD_BYTES:
                raise HTTPException(status_code=413, detail="El archivo supera el límite de 10 MB")
            wb = load_workbook(BytesIO(data), data_only=True)
            filas = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    filas.append(" | ".join("" if v is None else str(v) for v in row))
            contenido_extraido = "\n".join(filas)
        else:
            return {"contexto": f"\n[Archivo adjunto detectado: {file.filename} (Formato binario no procesado como texto plano)].\n"}

        if not contenido_extraido.strip():
            return {"contexto": ""}

        contexto_final = f"\n\n--- INICIO ARCHIVO ADJUNTO ({file.filename}) ---\n{contenido_extraido}\n--- FIN ARCHIVO ADJUNTO ---"
        return {"contexto": contexto_final}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")


# ==============================================================================
#  MÓDULO: CALCULADORA DE AMPLIFICADORES + SVG + BIAS (con validación de división por cero)
# ==============================================================================

def _formato_resistencia(valor_ohm: float) -> str:
    if valor_ohm >= 1e6:
        return f"{round(valor_ohm/1e6, 2)} MΩ"
    elif valor_ohm >= 1000:
        return f"{round(valor_ohm/1000, 2)} kΩ"
    else:
        return f"{round(valor_ohm, 1)} Ω"


def _formato_capacitor(valor_f: float) -> str:
    if valor_f >= 1e-3:
        return f"{round(valor_f*1e3, 2)} mF"
    elif valor_f >= 1e-6:
        return f"{round(valor_f*1e6, 1)} µF"
    elif valor_f >= 1e-9:
        return f"{round(valor_f*1e9, 1)} nF"
    else:
        return f"{round(valor_f*1e12, 1)} pF"


def calcular_clase_a(vcc: float, p_out: float, rl: float, transistor: str) -> dict:
    """Amplificador Clase A con transistor BJT en emisor común."""
    if rl <= 0 or vcc <= 0:
        raise HTTPException(status_code=400, detail="Vcc y carga deben ser mayores a cero.")
    
    vce_q = vcc / 2
    ic_q = vcc / (2 * rl)
    if ic_q <= 0:
        raise HTTPException(status_code=400, detail="La corriente de colector es cero o negativa. Ajuste Vcc o RL.")
    
    p_max_teorica = (vcc ** 2) / (8 * rl)
    eficiencia = 25.0
    
    ve = vcc * 0.1
    re = ve / ic_q
    vb = 0.7 + ve
    
    hfe = 50
    ib = ic_q / hfe
    i_divisor = 10 * ib
    r1_r2_total = vcc / i_divisor
    r2 = vb / i_divisor
    r1 = r1_r2_total - r2
    
    rc = (vcc - vce_q - ve) / ic_q
    ce = 1 / (2 * math.pi * 20 * (re / 11))
    cc = 1 / (2 * math.pi * 20 * rl)

    return {
        "topologia": "Clase A — Emisor Común",
        "transistor": transistor,
        "punto_bias": {
            "Vce_Q": round(vce_q, 2),
            "Ic_Q_mA": round(ic_q * 1000, 1),
            "Ib_Q_uA": round(ib * 1e6, 1),
            "Pd_transistor_W": round(vce_q * ic_q, 2),
        },
        "componentes": {
            "R1": _formato_resistencia(r1),
            "R2": _formato_resistencia(r2),
            "Rc": _formato_resistencia(rc),
            "Re": _formato_resistencia(re),
            "Ce": _formato_capacitor(ce),
            "Cc_entrada": _formato_capacitor(cc),
            "Cc_salida": _formato_capacitor(cc),
        },
        "performance": {
            "Potencia_max_W": round(p_max_teorica, 2),
            "Eficiencia_pct": eficiencia,
            "Vcc_V": vcc,
            "Rl_ohm": rl,
        },
        "notas": [
            f"Transistor {transistor}: verificar Ic_max > {round(ic_q*3*1000,0)}mA y Vce_max > {vcc}V",
            f"Pd disipada en transistor: {round(vce_q*ic_q,2)}W — usar disipador adecuado",
            "Clase A: alta linealidad, baja eficiencia. Ideal para etapas de pequeña señal.",
        ]
    }


def calcular_clase_ab(vcc: float, p_out: float, rl: float, transistor: str) -> dict:
    """Amplificador Clase AB push-pull complementario."""
    if rl <= 0 or p_out <= 0:
        raise HTTPException(status_code=400, detail="Potencia y carga deben ser mayores a cero.")
    
    v_pico = math.sqrt(2 * p_out * rl)
    vsat = 2.0
    vcc_minimo = v_pico + vsat
    vcc_usado = max(vcc, vcc_minimo)
    
    i_pico = v_pico / rl
    iq = i_pico * 0.07
    if iq <= 0:
        raise HTTPException(status_code=400, detail="Corriente de polarización nula. Ajuste la potencia o carga.")
    
    re = 0.33 if i_pico > 1 else 1.0
    v_bias = 1.4
    r_bias = v_bias / (iq * 10)
    eficiencia = 65.0
    
    pd_total = (vcc_usado * i_pico / math.pi) - p_out
    pd_por_transistor = max(0, pd_total / 2)

    return {
        "topologia": "Clase AB — Push-Pull Complementario",
        "transistor": f"{transistor} (NPN) + complementario PNP",
        "punto_bias": {
            "Vce_Q": round(vcc_usado / 2, 2),
            "Iq_mA": round(iq * 1000, 1),
            "Ipico_A": round(i_pico, 2),
            "Vpico_salida_V": round(v_pico, 2),
            "Pd_por_transistor_W": round(pd_por_transistor, 2),
        },
        "componentes": {
            "Re_NPN": _formato_resistencia(re),
            "Re_PNP": _formato_resistencia(re),
            "R_bias": _formato_resistencia(r_bias),
            "D_bias_1": "1N4148 o Vbe multiplier",
            "D_bias_2": "1N4148 o Vbe multiplier",
            "Cc_salida": _formato_capacitor(1 / (2 * math.pi * 20 * rl)),
        },
        "performance": {
            "Potencia_salida_W": round(p_out, 2),
            "Potencia_max_W": round((vcc_usado ** 2) / (2 * rl), 2),
            "Eficiencia_pct": eficiencia,
            "Vcc_V": vcc_usado,
            "Rl_ohm": rl,
        },
        "notas": [
            f"Vcc mínimo recomendado para {p_out}W en {rl}Ω: {round(vcc_minimo,1)}V",
            f"Cada transistor debe disipar al menos {round(pd_por_transistor,1)}W — usar disipador",
            "Agregar capacitor de compensación en la realimentación para estabilidad",
            "El Vbe multiplier es preferible a diodos para ajuste fino del bias",
        ]
    }


def generar_svg_clase_a(datos: dict) -> str:
    comp = datos["componentes"]
    bias = datos["punto_bias"]
    perf = datos["performance"]
    vcc = perf["Vcc_V"]

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 520 480" style="background:#1a1a2e;border-radius:12px;font-family:monospace">
  <defs>
    <marker id="arr" markerWidth="6" markerHeight="6" refX="3" refY="3" orient="auto">
      <path d="M0,0 L6,3 L0,6 Z" fill="#10b981"/>
    </marker>
  </defs>

  <text x="260" y="30" text-anchor="middle" fill="#10b981" font-size="15" font-weight="bold">Amplificador Clase A — Emisor Común</text>
  <text x="260" y="48" text-anchor="middle" fill="#6b7280" font-size="11">Vcc={vcc}V | Ic={bias["Ic_Q_mA"]}mA | Vce={bias["Vce_Q"]}V</text>

  <line x1="260" y1="65" x2="260" y2="80" stroke="#ef4444" stroke-width="2"/>
  <text x="260" y="62" text-anchor="middle" fill="#ef4444" font-size="12" font-weight="bold">+Vcc ({vcc}V)</text>

  <line x1="260" y1="80" x2="260" y2="110" stroke="#10b981" stroke-width="1.5"/>
  <rect x="245" y="110" width="30" height="50" rx="3" fill="none" stroke="#10b981" stroke-width="1.5"/>
  <line x1="260" y1="160" x2="260" y2="190" stroke="#10b981" stroke-width="1.5"/>
  <text x="245" y="142" text-anchor="end" fill="#f3f4f6" font-size="11">Rc</text>
  <text x="243" y="155" text-anchor="end" fill="#fbbf24" font-size="10">{comp["Rc"]}</text>

  <circle cx="260" cy="215" r="28" fill="none" stroke="#3b82f6" stroke-width="2"/>
  <line x1="260" y1="190" x2="260" y2="203" stroke="#10b981" stroke-width="1.5"/>
  <line x1="232" y1="215" x2="248" y2="215" stroke="#10b981" stroke-width="2"/>
  <line x1="248" y1="203" x2="260" y2="203" stroke="#10b981" stroke-width="1.5"/>
  <line x1="248" y1="227" x2="260" y2="227" stroke="#10b981" stroke-width="1.5"/>
  <line x1="248" y1="203" x2="248" y2="227" stroke="#10b981" stroke-width="2"/>
  <polygon points="255,224 265,231 255,234" fill="#10b981"/>
  <text x="295" y="210" fill="#93c5fd" font-size="12" font-weight="bold">{datos["transistor"]}</text>
  <text x="295" y="223" fill="#6b7280" font-size="10">BJT NPN</text>

  <line x1="260" y1="243" x2="260" y2="265" stroke="#10b981" stroke-width="1.5"/>
  <rect x="245" y="265" width="30" height="40" rx="3" fill="none" stroke="#10b981" stroke-width="1.5"/>
  <line x1="260" y1="305" x2="260" y2="330" stroke="#10b981" stroke-width="1.5"/>
  <text x="245" y="282" text-anchor="end" fill="#f3f4f6" font-size="11">Re</text>
  <text x="243" y="295" text-anchor="end" fill="#fbbf24" font-size="10">{comp["Re"]}</text>

  <line x1="260" y1="330" x2="260" y2="345" stroke="#10b981" stroke-width="1.5"/>
  <line x1="240" y1="345" x2="280" y2="345" stroke="#10b981" stroke-width="2"/>
  <line x1="247" y1="351" x2="273" y2="351" stroke="#10b981" stroke-width="1.5"/>
  <line x1="254" y1="357" x2="266" y2="357" stroke="#10b981" stroke-width="1"/>
  <text x="260" y="370" text-anchor="middle" fill="#6b7280" font-size="10">GND</text>

  <line x1="260" y1="290" x2="310" y2="290" stroke="#10b981" stroke-width="1" stroke-dasharray="3,2"/>
  <line x1="310" y1="275" x2="310" y2="305" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="315" y1="275" x2="315" y2="305" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="315" y1="290" x2="340" y2="290" stroke="#10b981" stroke-width="1" stroke-dasharray="3,2"/>
  <line x1="340" y1="290" x2="340" y2="345" stroke="#10b981" stroke-width="1" stroke-dasharray="3,2"/>
  <text x="312" y="270" text-anchor="middle" fill="#f3f4f6" font-size="10">Ce</text>
  <text x="312" y="260" text-anchor="middle" fill="#fbbf24" font-size="10">{comp["Ce"]}</text>

  <line x1="140" y1="80" x2="260" y2="80" stroke="#ef4444" stroke-width="1.5"/>
  <line x1="140" y1="80" x2="140" y2="130" stroke="#10b981" stroke-width="1.5"/>
  <rect x="125" y="130" width="30" height="45" rx="3" fill="none" stroke="#10b981" stroke-width="1.5"/>
  <line x1="140" y1="175" x2="140" y2="215" stroke="#10b981" stroke-width="1.5"/>
  <text x="123" y="150" text-anchor="end" fill="#f3f4f6" font-size="11">R1</text>
  <text x="121" y="163" text-anchor="end" fill="#fbbf24" font-size="10">{comp["R1"]}</text>

  <rect x="125" y="215" width="30" height="45" rx="3" fill="none" stroke="#10b981" stroke-width="1.5"/>
  <line x1="140" y1="260" x2="140" y2="345" stroke="#10b981" stroke-width="1.5"/>
  <text x="123" y="235" text-anchor="end" fill="#f3f4f6" font-size="11">R2</text>
  <text x="121" y="248" text-anchor="end" fill="#fbbf24" font-size="10">{comp["R2"]}</text>

  <line x1="140" y1="215" x2="232" y2="215" stroke="#10b981" stroke-width="1.5"/>

  <line x1="60" y1="215" x2="100" y2="215" stroke="#10b981" stroke-width="1.5"/>
  <line x1="100" y1="200" x2="100" y2="230" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="105" y1="200" x2="105" y2="230" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="105" y1="215" x2="125" y2="215" stroke="#10b981" stroke-width="1.5"/>
  <text x="50" y="210" text-anchor="middle" fill="#a78bfa" font-size="11">IN</text>
  <text x="102" y="198" text-anchor="middle" fill="#f3f4f6" font-size="10">Cc</text>
  <text x="102" y="188" text-anchor="middle" fill="#fbbf24" font-size="10">{comp["Cc_entrada"]}</text>

  <line x1="260" y1="190" x2="370" y2="190" stroke="#10b981" stroke-width="1" stroke-dasharray="3,2"/>
  <line x1="370" y1="175" x2="370" y2="205" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="375" y1="175" x2="375" y2="205" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="375" y1="190" x2="430" y2="190" stroke="#10b981" stroke-width="1.5"/>
  <line x1="430" y1="190" x2="430" y2="230" stroke="#10b981" stroke-width="1.5"/>
  <rect x="415" y="230" width="30" height="20" rx="2" fill="none" stroke="#a78bfa" stroke-width="1.5"/>
  <text x="430" y="243" text-anchor="middle" fill="#a78bfa" font-size="10">RL</text>
  <text x="430" y="255" text-anchor="middle" fill="#fbbf24" font-size="10">{perf["Rl_ohm"]}Ω</text>
  <line x1="430" y1="250" x2="430" y2="345" stroke="#10b981" stroke-width="1.5"/>
  <text x="372" y="173" text-anchor="middle" fill="#f3f4f6" font-size="10">Cc</text>
  <text x="372" y="163" text-anchor="middle" fill="#fbbf24" font-size="10">{comp["Cc_salida"]}</text>
  <text x="470" y="195" fill="#a78bfa" font-size="11">OUT</text>

  <line x1="140" y1="345" x2="430" y2="345" stroke="#10b981" stroke-width="1.5"/>
</svg>'''
    return svg


def generar_svg_clase_ab(datos: dict) -> str:
    comp = datos["componentes"]
    bias = datos["punto_bias"]
    perf = datos["performance"]
    vcc = perf["Vcc_V"]

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 520 500" style="background:#1a1a2e;border-radius:12px;font-family:monospace">
  <text x="260" y="28" text-anchor="middle" fill="#10b981" font-size="15" font-weight="bold">Amplificador Clase AB — Push-Pull</text>
  <text x="260" y="46" text-anchor="middle" fill="#6b7280" font-size="11">±{round(vcc/2,1)}V | Ipico={bias["Ipico_A"]}A | P={perf["Potencia_salida_W"]}W</text>

  <text x="260" y="72" text-anchor="middle" fill="#ef4444" font-size="12" font-weight="bold">+Vcc ({round(vcc/2,1)}V)</text>
  <line x1="260" y1="75" x2="260" y2="105" stroke="#ef4444" stroke-width="2"/>

  <circle cx="260" cy="135" r="28" fill="none" stroke="#3b82f6" stroke-width="2"/>
  <line x1="248" y1="123" x2="248" y2="147" stroke="#10b981" stroke-width="2"/>
  <line x1="232" y1="132" x2="248" y2="132" stroke="#10b981" stroke-width="1.5"/>
  <line x1="248" y1="123" x2="265" y2="110" stroke="#10b981" stroke-width="1.5"/>
  <line x1="248" y1="147" x2="265" y2="160" stroke="#10b981" stroke-width="1.5"/>
  <polygon points="260,157 270,163 262,168" fill="#10b981"/>
  <text x="296" y="130" fill="#93c5fd" font-size="11" font-weight="bold">{datos["transistor"].split("(")[0].strip()}</text>
  <text x="296" y="143" fill="#6b7280" font-size="10">NPN</text>

  <line x1="265" y1="160" x2="265" y2="185" stroke="#10b981" stroke-width="1.5"/>
  <rect x="250" y="185" width="30" height="30" rx="3" fill="none" stroke="#10b981" stroke-width="1.5"/>
  <text x="248" y="198" text-anchor="end" fill="#f3f4f6" font-size="10">Re</text>
  <text x="246" y="209" text-anchor="end" fill="#fbbf24" font-size="10">{comp["Re_NPN"]}</text>

  <line x1="265" y1="215" x2="265" y2="250" stroke="#10b981" stroke-width="1.5"/>
  <circle cx="265" cy="250" r="4" fill="#10b981"/>
  <line x1="265" y1="250" x2="390" y2="250" stroke="#10b981" stroke-width="1.5"/>

  <line x1="390" y1="235" x2="390" y2="265" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="396" y1="235" x2="396" y2="265" stroke="#f3f4f6" stroke-width="2"/>
  <line x1="396" y1="250" x2="450" y2="250" stroke="#10b981" stroke-width="1.5"/>
  <rect x="435" y="240" width="30" height="20" rx="2" fill="none" stroke="#a78bfa" stroke-width="1.5"/>
  <text x="450" y="235" fill="#a78bfa" font-size="11">RL</text>
  <text x="450" y="275" fill="#fbbf24" font-size="10">{perf["Rl_ohm"]}Ω</text>
  <line x1="450" y1="260" x2="450" y2="350" stroke="#10b981" stroke-width="1.5"/>

  <rect x="250" y="285" width="30" height="30" rx="3" fill="none" stroke="#10b981" stroke-width="1.5"/>
  <line x1="265" y1="250" x2="265" y2="285" stroke="#10b981" stroke-width="1.5"/>
  <text x="248" y="298" text-anchor="end" fill="#f3f4f6" font-size="10">Re</text>
  <text x="246" y="309" text-anchor="end" fill="#fbbf24" font-size="10">{comp["Re_PNP"]}</text>

  <line x1="265" y1="315" x2="265" y2="340" stroke="#10b981" stroke-width="1.5"/>
  <circle cx="260" cy="365" r="28" fill="none" stroke="#3b82f6" stroke-width="2"/>
  <line x1="248" y1="353" x2="248" y2="377" stroke="#10b981" stroke-width="2"/>
  <line x1="232" y1="365" x2="248" y2="365" stroke="#10b981" stroke-width="1.5"/>
  <line x1="248" y1="353" x2="265" y2="340" stroke="#10b981" stroke-width="1.5"/>
  <line x1="248" y1="377" x2="265" y2="390" stroke="#10b981" stroke-width="1.5"/>
  <polygon points="260,343 270,337 262,332" fill="#10b981"/>
  <text x="296" y="365" fill="#93c5fd" font-size="11" font-weight="bold">Complementario</text>
  <text x="296" y="378" fill="#6b7280" font-size="10">PNP</text>

  <line x1="260" y1="393" x2="260" y2="425" stroke="#ef4444" stroke-width="2"/>
  <text x="260" y="440" text-anchor="middle" fill="#ef4444" font-size="12" font-weight="bold">-Vcc ({round(vcc/2,1)}V)</text>

  <line x1="450" y1="350" x2="450" y2="415" stroke="#10b981" stroke-width="2"/>
  <line x1="280" y1="415" x2="450" y2="415" stroke="#10b981" stroke-width="2"/>
  <line x1="300" y1="415" x2="300" y2="430" stroke="#10b981" stroke-width="2"/>
  <line x1="287" y1="431" x2="313" y2="431" stroke="#10b981" stroke-width="1.5"/>
  <line x1="294" y1="437" x2="306" y2="437" stroke="#10b981" stroke-width="1"/>
  <text x="300" y="450" text-anchor="middle" fill="#6b7280" font-size="10">GND</text>
</svg>'''
    return svg


def generar_svg_amplificador(datos: dict) -> str:
    top = datos.get("topologia", "")
    if "Clase AB" in top:
        return generar_svg_clase_ab(datos)
    elif "Clase A" in top:
        return generar_svg_clase_a(datos)
    return "<svg><text fill='white' x='10' y='20'>Topología no soportada aún</text></svg>"


def calcular_crossover(frecuencia_hz: float, impedancia_ohm: float, orden: int) -> dict:
    if orden == 1:
        capacitor = 1 / (2 * math.pi * frecuencia_hz * impedancia_ohm)
        inductor = impedancia_ohm / (2 * math.pi * frecuencia_hz)
        componentes = {"C_pasa_altos": _formato_capacitor(capacitor), "L_pasa_bajos": f"{round(inductor * 1000, 2)} mH"}
    else:
        capacitor = 1 / (2 * math.pi * frecuencia_hz * impedancia_ohm * math.sqrt(2))
        inductor = impedancia_ohm / (2 * math.pi * frecuencia_hz * math.sqrt(2))
        componentes = {"C_por_rama": _formato_capacitor(capacitor), "L_por_rama": f"{round(inductor * 1000, 2)} mH"}
    return {
        "herramienta": "Crossover pasivo",
        "componentes": componentes,
        "notas": [
            f"Frecuencia de corte: {round(frecuencia_hz, 1)} Hz",
            f"Impedancia nominal usada: {round(impedancia_ohm, 2)} Ω",
            "Usar componentes con tolerancia baja y potencia adecuada para audio.",
        ],
    }


def calcular_caja_sellada(volumen_litros: float, vas_litros: float, qts: float) -> dict:
    qtc = qts * math.sqrt((vas_litros / volumen_litros) + 1)
    relacion = vas_litros / ((qtc / qts) ** 2 - 1) if qtc > qts else volumen_litros
    return {
        "herramienta": "Caja sellada",
        "componentes": {"Qtc_estimado": round(qtc, 3), "Vb_confirmado": f"{round(volumen_litros, 2)} L", "Vb_para_mismo_Qtc": f"{round(relacion, 2)} L"},
        "notas": [
            "Qtc cercano a 0.707 suele ser una alineación Butterworth equilibrada.",
            "Valores mayores a 1.0 pueden sonar con pico/resonancia marcada.",
            "Verificar desplazamiento del driver, relleno interno y pérdidas reales de la caja.",
        ],
    }


def calcular_impedancia_paralelo(impedancias_ohm: List[float]) -> dict:
    impedancias = [z for z in impedancias_ohm if z > 0]
    if not impedancias:
        raise HTTPException(status_code=400, detail="Ingresá al menos una impedancia positiva")
    equivalente = 1 / sum(1 / z for z in impedancias)
    return {
        "herramienta": "Impedancia en paralelo",
        "componentes": {"Z_equivalente": f"{round(equivalente, 3)} Ω", "Cargas": ", ".join(f"{z} Ω" for z in impedancias)},
        "notas": [
            "Verificar que el amplificador soporte la impedancia equivalente calculada.",
            "En parlantes reales, la impedancia cambia con la frecuencia.",
        ],
    }


@app.post("/herramientas/audio")
async def calcular_herramienta_audio(req: AudioToolRequest):
    if req.herramienta == "crossover":
        if req.frecuencia_hz is None or req.impedancia_ohm is None:
            raise HTTPException(status_code=400, detail="Frecuencia e impedancia son obligatorias")
        return calcular_crossover(req.frecuencia_hz, req.impedancia_ohm, req.orden)
    if req.herramienta == "caja_sellada":
        if req.volumen_litros is None or req.vas_litros is None or req.qts is None:
            raise HTTPException(status_code=400, detail="Volumen, Vas y Qts son obligatorios")
        return calcular_caja_sellada(req.volumen_litros, req.vas_litros, req.qts)
    if req.herramienta == "impedancia_paralelo":
        return calcular_impedancia_paralelo(req.impedancias_ohm or [])
    raise HTTPException(status_code=400, detail="Herramienta no soportada")


# --- ENDPOINT PRINCIPAL ---

@app.post("/amplificador")
async def calcular_amplificador(req: AmpRequest):
    try:
        if req.topologia in ("clase_a", "common_emitter"):
            datos = calcular_clase_a(req.vcc, req.potencia_w, req.impedancia_carga, req.transistor)
        elif req.topologia == "clase_ab":
            datos = calcular_clase_ab(req.vcc, req.potencia_w, req.impedancia_carga, req.transistor)
        else:
            raise HTTPException(status_code=400, detail="Topología no soportada")
        
        svg_render = generar_svg_amplificador(datos)
        datos["svg"] = svg_render
        
        return JSONResponse(content=datos)
    except HTTPException:
        raise
    except ZeroDivisionError as e:
        raise HTTPException(status_code=400, detail=f"División por cero en el cálculo: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en cálculo: {str(e)}")


if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)